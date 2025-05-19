import imgkit
import os
import excel2img
from subprocess import run
from pdf2image import convert_from_path
import markdown
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv

def csv2excel2(csv_file):
    wb = Workbook()
    ws = wb.active

    # 读取CSV文件并写入到Excel
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row_index, row in enumerate(reader, start=1):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)

    # 自动调整列宽和行高
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 保存为Excel文件
    file_name = os.path.splitext(csv_file)[0] + '.xlsx'
    # print(file_name)
    wb.save(file_name)
    # print('转换成功')
    return file_name

def csv2excel(csv_file):
    data = pd.read_csv(csv_file, header=None)
    excel_file = csv_file.split('.')[0] + '.xlsx'  # 替换为你希望保存的Excel文件名
    data.to_excel(excel_file, index=False, header=False)
    print(f'已成功将{csv_file}转换为{excel_file}')
    return excel_file

def latex2png(latex_table_code, output_image):
    # latex_table_code: latex格式的表格代码
    # output_image: 输出的图片文件名，例如 "output.png"
    full_latex_code = r"""
    \documentclass{standalone}
    \usepackage[a4paper, margin=1in]{geometry}
    \usepackage{multirow}  % 允许多行合并单元格
    \usepackage{amsmath}   % 数学符号支持
    \usepackage{booktabs}  % 美化表格线
    \usepackage{graphicx}
    \usepackage[utf8]{inputenc}
    \usepackage{ctex}  % 处理中文字符


    \usepackage{booktabs}
    \begin{document}
    """ + latex_table_code + r"""
    \end{document}
    """

    # 将完整的 LaTeX 文档写入文件
    with open("table.tex", "w") as f:
        f.write(full_latex_code)

    # 使用 pdflatex 将 LaTeX 文件编译为 PDF
    run([r"D:\LaTex\miktex\bin\x64\pdflatex.exe", "table.tex"], check=True)

    # 清理辅助文件
    os.remove("table.aux")
    os.remove("table.log")

    images = convert_from_path('table.pdf', poppler_path=r"D:\poppler\poppler-24.07.0\Library\bin")
    
    # 支持多页的表格形式
    for i, image in enumerate(images):
        image.save(f"{output_image[:-4]}.png", "PNG")
    
    print(f"已成功将 LaTeX 表格转换为图片 {output_image}")
        
def table2img(table_path, output_image, sheet_name):
    xlsx_path = 'aaaaa'  # 初始化 xlsx_path 变量
    
    if table_path.endswith('.csv'):
        xlsx_path = csv2excel2(table_path)
    else:
        xlsx_path = table_path
    
    print(xlsx_path)
    # sheet_list = table_path.split('.')[0].split('\\')[-1] + '.html'
    try:
        excel2img.export_img(xlsx_path, output_image, sheet_name, None)
        os.remove(xlsx_path)
        print("表格已成功转换为图片")
    except Exception as e:
        print("图片保存错误", e)
        
def html2img(html_path, output_image):
    options = {
        'encoding': "UTF-8",  # 确保使用 UTF-8 编码
    }

    config = imgkit.config(wkhtmltoimage=r'D:\wkhtmltox\wkhtmltopdf\bin\wkhtmltoimage.exe')

    # 将HTML文件转换为图片
    imgkit.from_file(html_path, output_image, config=config, options=options)
    
def markdown2img(markdown_path, output_image):
    with open(markdown_path, 'r', encoding='utf-8') as md_file:
        md_content = md_file.read()

# 将 markdown 转换为 HTML
    html_content = markdown.markdown(md_content, extensions=['tables'])

    style = """
    <style>
        html, body {
            margin: 0;
            padding: 0;
            height: 100%;  /* 设置 html 和 body 的高度 */
        }
        body {
            display: flex;
            justify-content: center;  /* 水平居中 */
            align-items: center;      /* 垂直居中 */
            min-height: 100%;         /* 保证 body 高度至少为 100% */
            overflow: hidden;         /* 禁止溢出内容产生的滚动条 */
        }
        table {
            width: 100%;  /* 表格宽度设置为页面的 80% */
            height: 100%;
            border: 1px solid black;
            border-collapse: collapse;
        }
        th, td {
            border: 1px solid black;
            padding: 8px;
            text-align: left;
        }
    </style>
    """

    html_with_style = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Markdown Table</title>
        {style}
    </head>
    <body>
    {html_content}
    </body>
    </html>
    """

    # 将生成的 HTML 写入文件
    with open('output.html', 'w', encoding='utf-8') as html_file:
        html_file.write(html_with_style)

    options = {
        'encoding': "UTF-8",  # 确保使用 UTF-8 编码
    }
    config = imgkit.config(wkhtmltoimage=r'D:\wkhtmltox\wkhtmltopdf\bin\wkhtmltoimage.exe')
    # 将HTML文件转换为图片
    imgkit.from_file('output.html', output_image, config=config, options=options)
    os.remove('output.html')

    print('markdown表格已成功转换为图片。')
    
    
if __name__ == '__main__':
    
    '''data in latex format is given in json file where each data is a dictionary with key *latex_code* '''
    
    # latex_json = './latex/small.json'
    # with open(latex_json, 'r', encoding='utf-8') as f:
    #     latex_data = json.load(f)
    # for i, latex in enumerate(latex_data):
    #     print(i)
    #     latex_code = latex['latex_code']
    #     latex2png(latex_code, f'latex_imgs/{i}.png')
        
    '''data in html format is given in html file where each data is a single file'''
    
    html_folder = r'D:\Paper\VLM\language_repo\language\totto\train_data'
    html_files = os.listdir(html_folder)
    for i, html_file in enumerate(html_files):
        html_file_path = os.path.join(html_folder, html_file)
        html2img(html_file_path, f'html_imgs/{i}.png')
        
    '''data in csv format is given in csv file where each data is a single file'''
    
    # table_folder = r'D:\Edge_Download\Table-Fact-Checking-master\data\processed_csv'
    # table_files = os.listdir(table_folder)
    # for i, table_file in enumerate(table_files):
    #     table_file_path = os.path.join(table_folder, table_file)
    #     print(i, '         ', table_file_path)
    #     if table_file_path.endswith('.csv'):
    #         sheet_name = 'Sheet'
    #     else:
    #         sheet_name = table_file.split('.')[0].split('\\')[-1] + '.html'
    #     table2img(table_file_path, f'csv_imgs/{i}.png', sheet_name)
        
    